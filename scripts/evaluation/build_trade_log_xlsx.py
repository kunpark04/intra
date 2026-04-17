"""Build `evaluation/top_trade_log.xlsx` — a pure, human-readable trade log.

Two sheets:
  - `unfiltered`   : all raw trades across the C1-selected top-10 (chronological)
  - `ml2_filtered` : event-driven portfolio of V3-filtered trades (fixed $500)

Rows are shaded green (win) / red (loss) for quick scanning. Header is bold +
frozen; column widths auto-sized to content.

Run from repo root:
    python scripts/evaluation/build_trade_log_xlsx.py
"""
from __future__ import annotations

import importlib.util
import json
import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(REPO))

from scripts.evaluation.composed_strategy_runner import (  # noqa: E402
    load_test_bars,
    run_strategy,
)

_spec = importlib.util.spec_from_file_location(
    "_v3eval", REPO / "scripts/evaluation/final_holdout_eval_v3_c1_fixed500.py"
)
v3eval = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(v3eval)

TOP_STRATEGIES = REPO / "evaluation" / "top_strategies.json"
OUT_XLSX = REPO / "evaluation" / "top_trade_log.xlsx"
FIXED_DOLLARS = 500.0

GREEN = PatternFill("solid", start_color="C8E6C9", end_color="C8E6C9")
RED = PatternFill("solid", start_color="FFCDD2", end_color="FFCDD2")
HEADER_FONT = Font(bold=True)


def build_unfiltered(strategies: list[dict], bars: pd.DataFrame) -> pd.DataFrame:
    """Run composed_strategy_runner per combo, concatenate, sort by date."""
    frames = []
    for s in strategies:
        gcid = s["global_combo_id"]
        print(f"  [unfiltered] {gcid}...", flush=True)
        r = run_strategy(s, bars=bars)
        if r["trades"].empty:
            continue
        t = r["trades"].copy()
        t.insert(0, "combo_id", gcid)
        frames.append(t)
    if not frames:
        return pd.DataFrame()
    df = (pd.concat(frames, ignore_index=True)
          .sort_values("date", kind="mergesort")
          .reset_index(drop=True))
    return df


def build_ml2(strategies: list[dict], bars: pd.DataFrame) -> pd.DataFrame:
    """Run the V3 stack then portfolio-simulate with fixed-$500 sizing.

    Returns one row per executed trade (event-driven, bar-overlap arbitration).
    """
    import lightgbm as lgb

    print("  Loading V3 booster + calibrators...", flush=True)
    booster = lgb.Booster(model_file=str(v3eval.v3inf.V3_BOOSTER))
    simple_cals = v3eval.v3inf._load_calibrators()
    two_stage = v3eval.v3inf._load_per_combo_calibrators()

    combos = []
    for s in strategies:
        gcid = s["global_combo_id"]
        print(f"  [ml2] {gcid}...", flush=True)
        try:
            c = v3eval.build_combo_trades_test(gcid, booster, simple_cals, two_stage)
        except Exception as e:
            c = {"combo_id": gcid, "error": str(e)}
        combos.append(c)

    events = []
    for ci, c in enumerate(combos):
        if c.get("error") or c.get("n_trades", 0) == 0:
            continue
        for ti in range(c["n_trades"]):
            events.append((int(c["entry_bar"][ti]), 0, ci, ti))
            events.append((int(c["exit_bar"][ti]), 1, ci, ti))
    events.sort(key=lambda e: (e[0], -e[1]))

    open_pos: dict[tuple[int, int], tuple[int, pd.Timestamp]] = {}
    rows = []
    bar_times = pd.to_datetime(bars["time"].to_numpy())
    for bar, kind, ci, ti in events:
        c = combos[ci]
        if kind == 0:
            contracts = int(FIXED_DOLLARS // (c["sl_pts"][ti] * v3eval.DOLLARS_PER_POINT))
            if contracts <= 0:
                continue
            open_pos[(ci, ti)] = (contracts, bar_times[bar])
        else:
            key = (ci, ti)
            if key not in open_pos:
                continue
            contracts, entry_time = open_pos.pop(key)
            sl_pts = float(c["sl_pts"][ti])
            pnl_pts = float(c["pnl_pts"][ti])
            pnl = pnl_pts * contracts * v3eval.DOLLARS_PER_POINT
            dollar_risk = sl_pts * contracts * v3eval.DOLLARS_PER_POINT
            rr = float(c["rr"])
            dollar_reward = dollar_risk * rr
            rows.append({
                "combo_id": c["combo_id"],
                "entry_time": entry_time,
                "exit_time": bar_times[bar] if bar < len(bar_times) else pd.NaT,
                "contracts": contracts,
                "sl_pts": round(sl_pts, 3),
                "rr": round(rr, 2),
                "dollar_risk": round(dollar_risk, 2),
                "dollar_reward": round(dollar_reward, 2),
                "pnl_pts": round(pnl_pts, 3),
                "label_win": int(c["label_win"][ti]),
                "pwin_simple": round(float(c["pwin_simple"][ti]), 4),
                "actual_pnl": round(pnl, 2),
            })
    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values("entry_time", kind="mergesort").reset_index(drop=True)
    return df


def _format_sheet(ws, df: pd.DataFrame) -> None:
    """Write df to ws with header row, row-fill by pnl sign, and autofit widths."""
    if df.empty:
        ws.append(["(no trades)"])
        return

    display = df.copy()
    for col in display.columns:
        if pd.api.types.is_datetime64_any_dtype(display[col]):
            display[col] = display[col].dt.strftime("%Y-%m-%d %H:%M")

    cols = list(display.columns)
    ws.append(cols)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    pnl_col = "actual_pnl" if "actual_pnl" in cols else None
    for _, row in display.iterrows():
        values = [None if pd.isna(v) else v for v in row.to_list()]
        ws.append(values)
        if pnl_col is not None:
            v = row[pnl_col]
            if pd.notna(v):
                try:
                    fill = GREEN if float(v) > 0 else RED if float(v) < 0 else None
                except (ValueError, TypeError):
                    fill = None
                if fill is not None:
                    for cell in ws[ws.max_row]:
                        cell.fill = fill

    ws.freeze_panes = "A2"
    for i, col in enumerate(cols, start=1):
        max_len = max(
            len(str(col)),
            display[col].astype(str).map(len).max() if not display.empty else 0,
        )
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 40)


def main() -> None:
    payload = json.loads(TOP_STRATEGIES.read_text())
    strategies = payload["top"]
    print(f"Loaded {len(strategies)} strategies from {TOP_STRATEGIES.name}")

    bars = load_test_bars()
    print(f"Test partition: {len(bars):,} bars "
          f"{bars['time'].iloc[0]} -> {bars['time'].iloc[-1]}")

    print("\n=== Building unfiltered trade log ===")
    df_raw = build_unfiltered(strategies, bars)
    print(f"unfiltered: {len(df_raw):,} trades")

    print("\n=== Building ML2-filtered trade log ===")
    df_ml2 = build_ml2(strategies, bars)
    print(f"ml2_filtered: {len(df_ml2):,} trades")

    wb = Workbook()
    wb.remove(wb.active)
    ws1 = wb.create_sheet("unfiltered")
    _format_sheet(ws1, df_raw)
    ws2 = wb.create_sheet("ml2_filtered")
    _format_sheet(ws2, df_ml2)

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    size = OUT_XLSX.stat().st_size
    print(f"\nWrote {OUT_XLSX.relative_to(REPO)}  ({size:,} B)")


if __name__ == "__main__":
    main()
